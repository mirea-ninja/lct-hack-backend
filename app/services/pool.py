from __future__ import annotations

import asyncio
import secrets
from io import BytesIO
from tempfile import NamedTemporaryFile

import aiohttp
import openpyxl
import pandas as pd
from fastapi import UploadFile
from pydantic import UUID4
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import config
from app.models import ApartmentCreate, QueryCreate, QueryExport, QueryGet, SubQueryCreate
from app.repositories import QueryRepository
from app.services.query import QueryService
from app.storage import get_s3_client


async def send_file(file: bytes, filename: str) -> str:
    async with get_s3_client() as client:
        await client.put_object(
            Bucket=config.STORAGE_BUCKET_NAME,
            Key=f"{filename}",
            Body=file,
        )
    return f"{config.STORAGE_ENDPOINT}/{config.STORAGE_BUCKET_NAME}/{filename}"


class PoolService:
    @staticmethod
    def _rename_columns(df: pd.DataFrame) -> pd.DataFrame:
        return df.rename(
            columns={
                "Местоположение": "address",
                "Количество комнат": "rooms",
                "Сегмент (Новостройка, современное жилье, старый жилой фонд)": "segment",
                "Этажность дома": "floors",
                "Материал стен (Кипич, панель, монолит)": "walls",
                "Этаж расположения": "floor",
                "Площадь квартиры, кв.м": "apartment_area",
                "Площадь кухни, кв.м": "kitchen_area",
                "Наличие балкона/лоджии": "has_balcony",
                "Удаленность от станции метро, мин. пешком": "distance_to_metro",
                "Состояние (без отделки, муниципальный ремонт, с современная отделка)": "quality",
                "Состояние (без отделки, муниципальный ремонт, современная отделка)": "quality",
            }
        )

    @staticmethod
    def _split_by_rooms(df: pd.DataFrame) -> list[pd.DataFrame]:
        df.replace("Студия", 0, inplace=True)
        df.replace("Да", True, inplace=True)
        df.replace("Нет", False, inplace=True)
        df["has_balcony"] = df["has_balcony"].astype("bool")
        df["rooms"] = df["rooms"].astype("int32")

        df.sort_values(by="rooms", inplace=True)

        dfs = [df[df["rooms"] == i] for i in range(df["rooms"].max() + 1)]
        return [i for i in dfs if not i.empty]

    @staticmethod
    async def _convert_address(address: str) -> tuple[float, float]:
        url = "https://suggestions.dadata.ru/suggestions/api/4_1/rs/suggest/address"
        headers = {
            "Authorization": f"Token {config.BACKEND_DADATA_TOKEN}",
            "Content-Type": "application/json",
            "Accept": "application/json",
        }
        async with aiohttp.ClientSession() as session:
            async with session.post(
                url,
                headers=headers,
                json={"query": address},
            ) as response:
                if response.status != 200:
                    return -1, -1
                data = await response.json()
                if data["suggestions"]:
                    return (
                        data["suggestions"][0]["data"]["geo_lat"],
                        data["suggestions"][0]["data"]["geo_lon"],
                    )
                return -1, -1

    @staticmethod
    async def _convert_dfs_to_model(dfs: list[pd.DataFrame]) -> list[SubQueryCreate]:
        sub_queries = []
        for df in dfs:
            to_dict = df.to_dict(orient="records")
            apartments = [ApartmentCreate(**apartment) for apartment in to_dict]
            sub_queries.append(SubQueryCreate(input_apartments=apartments))
        return sub_queries

    @staticmethod
    async def _create_random_name() -> str:
        return secrets.token_hex(8)

    @staticmethod
    async def create(db: AsyncSession, user: UUID4, name: str, file: UploadFile) -> QueryGet:
        read_file = await file.read()
        filename = await PoolService._create_random_name()
        await send_file(file=read_file, filename=f"{filename}.xlsx")

        df = pd.read_excel(BytesIO(read_file))
        df = PoolService._rename_columns(df)
        dfs = PoolService._split_by_rooms(df)

        for i in range(len(dfs)):
            if dfs[i].empty:
                continue

            dfs[i].drop_duplicates(inplace=True)

            addresses = dfs[i]["address"].to_list()
            unique_addresses = list(set(addresses))

            tasks = [PoolService._convert_address(address) for address in unique_addresses]
            results = await asyncio.gather(*tasks)
            addresses_dict = dict(zip(unique_addresses, results))

            dfs[i]["lat"] = dfs[i]["address"].map(addresses_dict).apply(lambda x: x[0])
            dfs[i]["lon"] = dfs[i]["address"].map(addresses_dict).apply(lambda x: x[1])
            dfs[i] = dfs[i][(dfs[i]["lat"] != -1) & (dfs[i]["lon"] != -1)]

            if name is None:
                name = dfs[i]["address"].iloc[0]

        sub_queries = await PoolService._convert_dfs_to_model(dfs)

        query = QueryCreate(
            name=name,
            sub_queries=sub_queries,
            created_by=user,
            updated_by=user,
            input_file=f"{config.STORAGE_ENDPOINT}/{config.STORAGE_BUCKET_NAME}/{filename}.xlsx",
        )

        query_db = await QueryService.create(db=db, model=query)

        for _ in range(len(query_db.sub_queries)):
            query_db.sub_queries.sort(key=lambda x: x.input_apartments[0].rooms)

        return query_db

    @staticmethod
    def _create_excel_columns(ws, include_adjustments: bool):
        ws["A1"] = "Местоположение"
        ws["B1"] = "Количество комнат"
        ws["C1"] = "Сегмент"
        ws["D1"] = "Этажность дома"
        ws["E1"] = "Материал стен"
        ws["F1"] = "Этаж расположения"
        ws["G1"] = "Площадь квартиры, кв.м"
        ws["H1"] = "Площадь кухни, кв.м"
        ws["I1"] = "Наличие балкона/лоджии"
        ws["J1"] = "Удаленность от станции метро, мин. пешком"
        ws["K1"] = "Состояние"
        ws["L1"] = "Цена за кв.м"
        ws["M1"] = "Цена"

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["G"].width = 16
        ws.column_dimensions["H"].width = 16
        ws.column_dimensions["I"].width = 16
        ws.column_dimensions["J"].width = 16
        ws.column_dimensions["K"].width = 16
        ws.column_dimensions["L"].width = 16
        ws.column_dimensions["M"].width = 25

        if include_adjustments:
            ws["N1"] = "Цена за м2, %"
            ws["O1"] = "Этаж расположения, %"
            ws["P1"] = "Площадь квартиры м2, %"
            ws["Q1"] = "Площадь кухни м2, %"
            ws["R1"] = "Балкон/лоджия, %"
            ws["S1"] = "Удаленность от метро, %"
            ws["T1"] = "Состояние, %"

            ws.column_dimensions["N"].width = 16
            ws.column_dimensions["O"].width = 16
            ws.column_dimensions["P"].width = 16
            ws.column_dimensions["Q"].width = 16
            ws.column_dimensions["R"].width = 16
            ws.column_dimensions["S"].width = 16
            ws.column_dimensions["T"].width = 16

        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=20 if include_adjustments else 13):
            for cell in col:
                cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                cell.border = openpyxl.styles.Border(right=openpyxl.styles.Side(border_style="thin"))

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=20 if include_adjustments else 13):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    top=openpyxl.styles.Side(border_style="thick"), bottom=openpyxl.styles.Side(border_style="thick")
                )

    @staticmethod
    def _append_data_to_excel(ws, sub_query, include_adjustments: bool):
        for j in range(len(sub_query.input_apartments)):
            apartment = sub_query.input_apartments[j]
            row = [
                apartment.address,
                apartment.rooms,
                apartment.segment,
                apartment.floors,
                apartment.walls,
                apartment.floor,
                apartment.apartment_area,
                apartment.kitchen_area,
                "Да" if apartment.has_balcony else "Нет",
                apartment.distance_to_metro,
                apartment.quality,
                apartment.m2price,
                apartment.price,
            ]

            if include_adjustments and apartment.adjustment is not None:
                to_extend = [
                    f"{(apartment.adjustment.price_final - apartment.m2price) * 100 / apartment.m2price:.2f}%"
                    if apartment.m2price > 0
                    else "0%",
                    f"{apartment.adjustment.floor * 100:.2f}%",
                    f"{apartment.adjustment.apt_area * 100:.2f}%",
                    f"{apartment.adjustment.kitchen_area * 100:.2f}%",
                    f"{apartment.adjustment.has_balcony * 100:.2f}%",
                    f"{apartment.adjustment.distance_to_metro * 100:.2f}%",
                    f"{apartment.adjustment.quality}",
                ]

                to_extend = [f"+{x}" if x[0] != "-" else x for x in to_extend]
                row.extend(to_extend)

            ws.append(row)

    @staticmethod
    async def export(
        db: AsyncSession, guid: UUID4, include_adjustments: bool, split_by_lists: bool, user: UUID4
    ) -> QueryExport:
        query = await QueryService.get(db=db, guid=guid)
        wb = openpyxl.Workbook()
        ws = None

        if not split_by_lists:
            ws = wb.active
            ws.title = "Data"

            PoolService._create_excel_columns(ws, include_adjustments)

        for i in range(len(query.sub_queries)):
            sub_query = query.sub_queries[i]
            if split_by_lists:
                if i == 0:
                    ws = wb.active
                    ws.title = f"SubQuery {i+1}"
                ws = wb.create_sheet(f"SubQuery {i+1}")
                PoolService._create_excel_columns(ws, include_adjustments)

            PoolService._append_data_to_excel(ws, sub_query, include_adjustments)

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
        filename = await PoolService._create_random_name()
        await send_file(file=bytes(stream), filename=f"{filename}.xlsx")
        link = f"{config.STORAGE_ENDPOINT}/{config.STORAGE_BUCKET_NAME}/{filename}.xlsx"
        await QueryRepository.set_link(db=db, guid=guid, link=link)
        return QueryExport(link=link)
